import sqlite3 as sql
import openpyxl as xl
from pathlib import Path

#on peut faire de la ligne de commande avec click


def test_etat_reponse_appel(path_db, path_file):
    path_db = Path(path_db)
    path_file = Path(path_file)/Path("listeEtatsReponsesAppel.xlsx")

    try:
        con = sql.connect(path_db)
        file = xl.load_workbook(path_file, read_only=True)
        tab = file["Export Workbook"]

        rows = tab.rows
        next(rows)
        for line in rows:
            with con:
                con.execute("INSERT INTO etat_reponse VALUES(?,?)", (line[0].value, line[1].value))
        file.close()
    except sql.DatabaseError as e:
        print(e)
    finally:
        con.close()

def test_liste_ecoles(path_db, path_file):
    path_db = Path(path_db)
    path_file = Path(path_file)/Path("listeEcoles.xlsx")

    try:
        con = sql.connect(path_db)
        file = xl.load_workbook(path_file, read_only=True)
        tab = file["Export Workbook"]

        rows = tab.rows
        next(rows)
        for line in rows:
            with con:
                con.execute("INSERT INTO ecole VALUES(?,?)", (line[0].value, line[1].value))
        file.close()
    except sql.DatabaseError as e:
        print(e)
    finally:
        con.close()

def test_liste_etablissements(path_db, path_file):
    path_db = Path(path_db)
    path_file = Path(path_file)/Path("listeEtablissements.xlsx")

    try:
        con = sql.connect(path_db)
        file = xl.load_workbook(path_file, read_only=True)
        tab = file["Export Workbook"]

        rows = tab.rows
        next(rows)
        for line in rows:
            with con:
                con.execute("INSERT INTO etablissement VALUES(?,?,?,?,?,?)", (line[0].value, line[1].value, line[2].value, line[3].value, line[4].value, line[5].value))
        file.close()
    except sql.DatabaseError as e:
        print(e)
    finally:
        con.close()


def test_pays(path_db, path_file):
    path_db = Path(path_db)
    path_file = Path(path_file)/Path("Inscription.xlsx")

    try:
        con = sql.connect(path_db)
        file = xl.load_workbook(path_file, read_only=True)
        tab = file["Export Workbook"]
        L = []
        rows = tab.rows
        next(rows)
        for line in rows:
            with con:
                if line[7].value not in L:
                    L.append(line[7].value)
                    if isinstance(line[7].value, (int,float)) :
                        con.execute("INSERT INTO pays VALUES(?,?)", (line[7].value, line[8].value))
        file = xl.load_workbook(path_file, read_only=True)
        tab = file["Export Workbook"]
        rows = tab.rows
        next(rows)
        for line in rows:
            with con:
                if line[10].value not in L:
                    L.append(line[10].value)
                    if isinstance(line[10].value, (int,float)) :
                        con.execute("INSERT INTO pays VALUES(?,?)", (line[10].value, line[11].value))
        file.close()

    except sql.DatabaseError as e:
        print(e)
    finally:
        con.close()


def test_csp_parent(path_db, path_file):
    path_db = Path(path_db)
    path_file = Path(path_file)/Path("Inscription.xlsx")

    try:
        con = sql.connect(path_db)
        file = xl.load_workbook(path_file, read_only=True)
        tab = file["Export Workbook"]
        L = []
        rows = tab.rows
        next(rows)
        for line in rows:
            with con:
                if line[45].value not in L:
                    L.append(line[45].value)
                    if line[45].value.isdecimal():
                        con.execute("INSERT INTO csp_parent VALUES(?,?)", (int(line[45].value), line[46].value))
        file = xl.load_workbook(path_file, read_only=True)
        tab = file["Export Workbook"]
        rows = tab.rows
        next(rows)
        for line in rows:
            with con:
                if line[47].value not in L:
                    L.append(line[47].value)
                    if line[47].value.isdecimal() :
                        con.execute("INSERT INTO csp_parent VALUES(?,?)", (int(line[47].value), line[48].value))
        file.close()

    except sql.DatabaseError as e:
        print(e)
    finally:
        con.close()



if __name__=="__main__":
    #test_etat_reponse_appel("dataa.db", "../dow/data/public")
    #test_liste_ecoles("dataa.db","../dow/data/public")
    #test_liste_etablissements("dataa.db","../dow/data/public")
    #test_pays("dataa.db","../dow/data/public")
    #test_csp_parent("dataa.db","../dow/data/public")


