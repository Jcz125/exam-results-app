import pandas as pd
import sqlite3 as sql
import openpyxl as xl
from pathlib import Path

# dictionnaire classement
dico_class = {
    'ECRIT': 1,
    'ORAL': 2,
    'RANG_ADMISSIBLE': 3,
    'RANG_CLASSE': 4
    }

# dictionnaires des épreuves
dico_epreuves = [
    ##### ECRITES #####
    ("Langue", 28),
    ("Informatique ou Sciences industrielles", 599),
    ("Mathématiques 1", 600),
    ("Mathématiques 2", 601),
    ("Physique 1", 602),
    ("Physique 2", 603),
    ("Chimie", 604),
    ("Français", 605),
    ("Sciences industrielles", 606),
    ("Informatique", 1050),
    ("bonification_ecrit", 9898),
    ("total_ecrit", 9899),

    ##### SPECIFIQUES CMT #####
    ("QCM info/physique", 1),
    ("Entretien nouvelles technologies", 2),
    ("QCM anglais", 3),
    ("Mathématiques", 4),  # ajout de spe CMT

    ##### ORALES CMT #####
    ("Physique ou Sciences industrielles", 5),
    ("Entretien", 6),
    ("Anglais", 7),  # ajout de oral CMT
    ("Mathématiques", 8),  # ajout de oral CMT

    ##### ORALES CCMP #####
    ("Physique", 9),
    ("Français", 10),  # ajout de oral CCMP
    ("Anglais", 11),
    ("Mathématiques", 12),

    ##### TOTAL ORAL #####
    ("Mathématiques (harmonisé)", 400),
    ("Mathématiques (affichée)", 401),
    ("max_physique", 9900),
    ("max_anglais", 9901),
    ("bonification_oral", 9998),
    ("total_oral", 9999),

    ##### TOTAL (ECRIT+ORAL) ######
    ("total", 10000),

    ##### CLASSEMENT #####
    ("bonus_interclassement", 10198),
    ("total_avec_interclassement", 10199),
    ("note_entretien_exaequo", 10200),
    ("note_anglais_exaequo", 10201),

    ##### ECRITES PT #####
    ("Mathématiques B", 700),
    ("Mathématiques C", 701),
    ("Physique A", 702),
    ("Physique B", 703),
    ("Informatique modélisation", 704),
    ("Sciences industrielles A", 705),
    ("Français B", 706),
    ("Langue A", 707),

    ##### ECRITES TSI #####
    ("Mathématiques 1", 800),  # ajout de TSI
    ("Mathématiques 2", 801),  # ajout de TSI
    ("Physique 1", 802),  # ajout de TSI
    ("Physique 2", 803),  # ajout de TSI
    ("Français", 804),  # ajout de TSI
    ("Langue", 805),  # ajout de TSI
    ("Sciences industrielles", 806),  # ajout de TSI
    ("Informatique", 807),  # ajout de TSI

    ##### ORALES TSI #####
    ("Mathématiques 1", 13),  # ajout de oral TSI
    ("Mathématiques 2", 14),  # ajout de oral TSI
    ("Physique-chimie 1", 15),  # ajout de oral TSI
    ("Physique-chimie 2", 16),  # ajout de oral TSI
    ("Langue vivante", 17),  # ajout de oral TSI
    ("TP Physique-chimie", 18),  # ajout de oral TSI
    ("S2I", 19),  # ajout de oral TSI

    ##### ECRIT ATS #####
    ("Mathématiques", 956),
    ("Sc. Physiques", 957),
    ("Français", 958),
    ("Sciences Industrielles", 959),
    ("Anglais", 960),
    ("Moyenne", 9897),

    ##### ORALES ATS #####
    ("Mathématiques", 961),
    ("Sciences Physiques", 962),
    ("Génie électrique", 963),
    ("Génie mécanique", 964),
    ("Langue au choix (oral commun)", 981),
    ("Epreuve ATS", 1030),
    ("Moyenne", 9997),
    ]


liste_ecrits = [28, 599, 600, 601, 602, 603, 604, 605, 606, 1050, 9898, 9899, 700, 701, 702, 703, 704, 705, 706, 707,
    800, 801, 802, 803, 804, 805, 806, 807, 956, 957, 958, 959, 960, 9897]
liste_oraux = [5, 6, 7, 8, 9, 10, 11, 12, 400, 401, 9900, 9901, 9998, 9999, 400, 401, 9900, 9901, 9998, 9999, 13, 14,
    15, 16, 17, 18, 19, 961, 962, 963, 964, 981, 1030, 9997]
liste_specifique = [1, 2, 3, 4]
liste_classement = [10198, 10199, 10200, 10201]

liste_ep_MP = [28, 599, 600, 601, 602, 603, 604, 605, 1050, 9898, 9899, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 400, 401,
    9900, 9901, 9998, 9999, 10000, 10198, 10199, 10200, 10201]
liste_ep_PC = [28, 600, 601, 602, 603, 604, 605, 1050, 9898, 9899, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 400, 401,
    9900, 9901, 9998, 9999, 10000, 10198, 10199, 10200, 10201]
liste_ep_PSI = [28, 600, 601, 602, 603, 604, 605, 606, 1050, 9898, 9899, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 400,
    401, 9900, 9901, 9998, 9999, 10000, 10198, 10199, 10200, 10201]
liste_ep_PT = [700, 701, 702, 703, 704, 705, 706, 707, 9898, 9899, 1, 2, 3, 4, 5, 6, 7, 8, 400, 401, 9900, 9901, 9998,
    9999, 10000, 10198, 10199, 10200, 10201]
liste_ep_TSI = [800, 801, 802, 803, 804, 805, 806, 807, 9898, 9899, 1, 2, 3, 4, 13, 14, 15, 16, 17, 18, 19, 400, 401,
    9998, 9999, 10000, 10198, 10199, 10201]
liste_ep_ATS_ecrit = [9899, 9897, 9898, 956, 957, 958, 959, 960]
liste_ep_ATS_oral = [10000, 9997, 9998, 961, 962, 963, 964, 981, 1030]

#######import file inscription
def import_inscription(database: sql.Connection, path: Path):
    data = pd.read_excel(path/'Inscription.xlsx', header=1)

    c = database.cursor()

    for i in range(len(data)):

        ####### etat_dossier
        code_etat_dossier = data['CODE_ETAT_DOSSIER'][i]
        lib_etat_dossier = data['LIBELLE_ETAT_DOSSIER'][i]
        consigne_sql = "INSERT OR IGNORE INTO etat_dossier VALUES (?, ?);"
        c.execute(consigne_sql, (int(code_etat_dossier), lib_etat_dossier))

        ####### concours
        code_concours = data['CODE_CONCOURS'][i]
        lib_concours = data['LIBELLE_CONCOURS'][i]
        voie_concours = data['VOIE'][i]
        consigne_sql = "INSERT OR IGNORE INTO concours VALUES (?, ?, ?);"
        c.execute(consigne_sql, (int(code_concours), lib_concours, voie_concours))

        ####### autres_prenoms
        code_etudiant = data['CODE_CANDIDAT'][i]
        autre_prenom = data['AUTRES_PRENOMS'][i]
        consigne_sql = "INSERT OR IGNORE INTO autre_prenoms VALUES (?, ?);"
        c.execute(consigne_sql, (int(code_etudiant), autre_prenom))

        ####### serie_bac
        code_serie = data['CODE_SERIE'][i]
        nom_serie = data['SERIE'][i]
        consigne_sql = "INSERT OR IGNORE INTO serie_bac VALUES (?, ?);"
        c.execute(consigne_sql, (int(code_serie), nom_serie))

        ####### ep_option
        for j in range(1, 5):
            ep, op = data[f'EPREUVE{j}'][i], data[f'OPTION{j}'][i]
            consigne_sql = "INSERT OR IGNORE INTO ep_option(epreuve, option) VALUES (?, ?);"
            c.execute(consigne_sql, (ep, op))

    #########csp
        code = int(data['COD_CSP_PERE'][i])
        lib = data['LIB_CSP_PERE'][i]
        c.execute("INSERT OR IGNORE INTO csp_parent VALUES (?,?)", (code, lib))

        code = int(data['COD_CSP_MERE'][i])
        lib = data['LIB_CSP_MERE'][i]
        c.execute("INSERT OR IGNORE INTO csp_parent VALUES (?,?)", (code, lib))

    ##########pays
        code = int(data['CODE_PAYS_NAISSANCE'][i])
        pays = data['PAYS_NAISSANCE'][i]
        c.execute("INSERT OR IGNORE INTO pays(code, lib) VALUES (?,?)", (code, pays))

        code = int(data['CODE_PAYS_NATIONALITE'][i])
        nationalite = data['AUTRE_NATIONALITE'][i]
        lib = c.execute("SELECT lib FROM pays WHERE code = ?", (code,)).fetchone()
        if lib is not None:
            lib = lib[0]
        c.execute("INSERT OR REPLACE INTO pays VALUES (?,?,?)", (code, lib, nationalite))

#########candidat
    dico_etat_cl, dico_type_ad = extrait_etat_classes_type_admissible(path)
    for i in range(len(data)):
        code = int(data['CODE_CANDIDAT'][i])

        date_n = '-'.join(reversed(data['DATE_NAISSANCE'][i].split('/')))

        c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',
            (data['EPREUVE1'][i], data['OPTION1'][i]))
        o1 = c.fetchone()
        if o1 is not None:
            o1 = o1[0]

        c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',
            (data['EPREUVE2'][i], data['OPTION2'][i]))
        o2 = c.fetchone()
        if o2 is not None:
            o2 = o2[0]

        c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',
            (data['EPREUVE3'][i], data['OPTION3'][i]))
        o3 = c.fetchone()
        if o3 is not None:
            o3 = o3[0]

        c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',
            (data['EPREUVE4'][i], data['OPTION4'][i]))
        o4 = c.fetchone()
        if o4 is not None:
            o4 = o4[0]

        if data['DECLARATION_HANDICAP'][i] == 'non':
            handi = 0
        else:
            handi = 1

        #####classe en meme temps que candidat
        c.execute('SELECT code FROM classe WHERE lib = ?', (data['CLASSE'][i],))
        classe = c.fetchone()
        if classe is None:
            c.execute('INSERT INTO classe(lib) VALUES(?)', (data['CLASSE'][i],))
            c.execute('SELECT code FROM classe WHERE lib = ?', (data['CLASSE'][i],))
            classe = c.fetchone()
        classe = classe[0]

        c.execute(
            "INSERT INTO candidat VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (code, int(data['CIVILITE'][i]), data['NOM'][i], data['PRENOM'][i],
            date_n,
            classe, data['PUISSANCE'][i], int(data['CODE_CONCOURS'][i]), data['CODE_ETABLISSEMENT'][i],
            data['ADRESSE1'][i],
            data['ADRESSE2'][i], int(data['CP'][i]), data['VILLE'][i], int(data['CODE_PAYS'][i]),
            data['EMAIL'][i],
            data['TELEPHONE'][i], data['TEL_PORTABLE'][i], int(data['CODE_PAYS_NAISSANCE'][i]),data['VILLE_NAISSANCE'][i],
            int(data['CODE_PAYS_NATIONALITE'][i]), data['LIBELLE_VILLE_ECRIT'][i],
            data['NUMERO_INE'][i], int(data['COD_CSP_PERE'][i]), int(data['COD_CSP_MERE'][i]),
            int(data['CODE_SERIE'][i]), int(data['ANNEE_BAC'][i]),
            int(data['MOIS_BAC'][i]), data['MENTION'][i], data['CAN_DEP_BAC'][i],
            int(data['CODE_ETAT_DOSSIER'][i]), data['QUALITE'][i],
            handi, int(data['ARRONDISSEMENT_NAISSANCE'][i]), o1, o2, o3, o4, data['SUJET_TIPE'][i], dico_etat_cl.get(code, None),
            dico_type_ad.get(code,None)))

#########voeux
def import_voeux(database: sql.Connection, path: Path):
    for classe in ("MP", "TSI", "PT", "ATS", "PC", "PSI"):
        file = xl.load_workbook(path/f"listeVoeux_{classe}.xlsx", read_only=True)
        tab = file[file.sheetnames[0]]

        c = database.cursor()
        rows = tab.rows
        next(rows)
        for line in rows:
            c.execute("INSERT OR IGNORE INTO voeux VALUES (?,?,?)", (line[0].value, line[2].value, line[3].value))
    file.close()

#######ecole
def import_ecole(database: sql.Connection, path: Path):
    data = pd.read_excel(path/'listeEcoles.xlsx', header=0)
    c = database.cursor()

    for i in range(len(data)):
        nom = data['Nom _ecole'][i]
        code = int(data['Ecole'][i])

        c.execute("INSERT INTO ecole VALUES (?,?)", (code, nom))

#######etat_reponse
def import_etat_reponse(database: sql.Connection, path: Path):
    data = pd.read_excel(path/'listeEtatsReponsesAppel.xlsx', header=0)
    c = database.cursor()

    for i in range(len(data)):
        code = int(data['Ata _cod'][i])
        etat = data['Ata _lib'][i]

        c.execute("INSERT INTO etat_reponse VALUES (?,?)", (code, etat))

#######etablissement
def import_etablissement(database: sql.Connection, path: Path):
    data = pd.read_excel(path/'listeEtablissements.xlsx', header=0)
    c = database.cursor()

    for i in range(len(data)):
        rne = data['Rne'][i]
        type = data['Type _etab'][i]
        name = data['Etab'][i]
        cp = str(data['Code _postal _etab'][i])
        ville = data['Ville _etab'][i]
        pays = data['Pays _atab'][i]

        c.execute("INSERT INTO etablissement VALUES (?,?,?,?,?,?)", (rne, type, name, cp, ville, pays))

########candidat_ats
def import_ats(database: sql.Connection, path: Path):
    file = xl.load_workbook(path/"ADMISSIBLE_ATS.xlsx", read_only=True)
    tab = file[file.sheetnames[0]]

    c = database.cursor()
    rows = tab.rows
    C = {case.value: case.column - 1 for case in next(rows)}

    c.execute("INSERT INTO classe(lib) VALUES('ATS')")
    classe = c.execute("SELECT code FROM classe WHERE lib = 'ATS'").fetchone()
    classe = classe[0]

    for line in rows:
        if line[C['Civ _lib']].value == 'M.':
            civ = 1
        else:
            civ = 2

        pays = c.execute('SELECT code FROM pays WHERE lib = ?', (line[C['Can _pay _adr']].value,)).fetchone()
        if pays is None:
            c.execute('INSERT INTO pays(lib) VALUES(?)', (line[C['Can _pay _adr']].value,))
            pays = c.execute('SELECT code FROM pays WHERE lib = ?', (line[C['Can _pay _adr']].value,)).fetchone()
        pays = pays[0]

        c.execute(
            "INSERT INTO candidat(code, civ, nom, prenom, adresse1, adresse2, code_postal, commune, code_adr_pays, mail, tel, por, classe) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (line[C['Can _cod']].value, civ, line[C['Nom']].value, line[C['Prenom']].value, line[C['Can _ad 1']].value, line[C['Can _ad 2']].value, line[C['Can _cod _pos']].value, line[C['Can _com']].value, pays, line[C['Can _mel']].value, line[C['Can _tel']].value, line[C['Can _por']].value, classe))
    file.close()


####### epreuves
def import_epreuves(database: sql.Connection):
    c = database.cursor()

    for ep, cle in dico_epreuves:
        type = None

        if cle in liste_ecrits:
            type = "ECRIT"
        elif cle in liste_oraux:
            type = "ORAL"
        elif cle in liste_specifique:
            type = "SPECIFIQUE"
        elif cle in liste_classement:
            type = "CLASSEMENT"

        consigne_sql = "INSERT INTO epreuve VALUES (?, ?, ?)"
        c.execute(consigne_sql, (cle, ep, type))


##### notes
def import_notes(database: sql.Connection, path: Path):

    def func_aux(file, indices, liste_epreuve):
        c = database.cursor()
        tab = file[file.sheetnames[0]]

        rows = tab.rows
        next(rows)
        next(rows)
        for line in rows:
            for k in range(len(liste_epreuve)):
                if line[indices[k]].value != None and (liste_epreuve[k] > 9000 or 0.0 <= line[indices[k]].value <= 20.00):
                    consigne_sql = "INSERT INTO notes VALUES(?, ?, ?);"
                    c.execute(consigne_sql, (line[0].value, liste_epreuve[k], line[indices[k]].value))


    ### MP ###
    file = xl.load_workbook(path/"Classes_MP_CMT_spe_XXXX.xlsx", read_only=True)
    func_aux(file, list(range(13,24))+list(range(25,48)), liste_ep_MP)

    ### PC ###
    file = xl.load_workbook(path/"Classes_PC_CMT_spe_XXXX.xlsx", read_only=True)
    func_aux(file, list(range(13,23))+list(range(24,47)), liste_ep_PC)

    ### PSI ###
    file = xl.load_workbook(path/"Classes_PSI_CMT_spe_XXXX.xlsx", read_only=True)
    func_aux(file, list(range(13,24))+list(range(25,48)), liste_ep_PSI)

    ### PT ###
    file = xl.load_workbook(path/"Classes_PT_CMT_spe_XXXX.xlsx", read_only=True)
    func_aux(file, list(range(13,23))+list(range(24,43)), liste_ep_PT)

    ### TSI ###
    file = xl.load_workbook(path/"Classes_TSI_CMT_spe_XXXX.xlsx", read_only=True)
    func_aux(file, list(range(13,23))+list(range(24,43)), liste_ep_TSI)

    ### ATS ecrit ###
    file = xl.load_workbook(path/"ResultatEcrit_DD_MM_YYYY_ATS.xlsx", read_only=True)
    func_aux(file, list(range(2, 10)), liste_ep_ATS_ecrit)

    ### ATS oral ###
    file = xl.load_workbook(path/"ResultatOral_DD_MM_YYYY_ATS.xlsx", read_only=True)
    func_aux(file, [2, 3, 4, 6, 7, 8, 9, 10, 11], liste_ep_ATS_oral)


####### classement
def import_rang(database: sql.Connection, path: Path):
    c = database.cursor()
    def func_aux(data, type, rang, id):
        for i in range(len(data)):
            code_ed = data[id][i]
            rank = data[rang][i]
            if repr(rank) != 'nan':
                consigne_sql = "INSERT INTO classement(etudiant, rang, type) VALUES (?, ?, ?);"
                c.execute(consigne_sql, (int(code_ed), int(rank), dico_class[type]))


    ##### MP ECRIT #####
    ecrit = pd.read_excel(path/'Ecrit_MP.xlsx', header=0)
    func_aux(ecrit, 'ECRIT', 'rang', 'Can _cod')

    ##### PC ECRIT #####
    ecrit = pd.read_excel(path/'Ecrit_PC.xlsx', header=0)
    func_aux(ecrit, 'ECRIT', 'rang', 'Can _cod')

    ##### PSI ECRIT #####
    ecrit = pd.read_excel(path/'Ecrit_PSI.xlsx', header=0)
    func_aux(ecrit, 'ECRIT', 'rang', 'Can _cod')

    ##### PT ECRIT #####
    ecrit = pd.read_excel(path/'Ecrit_PT.xlsx', header=0)
    func_aux(ecrit, 'ECRIT', 'rang', 'Can _cod')

    ##### TSI ECRIT #####
    ecrit = pd.read_excel(path/'Ecrit_TSI.xlsx', header=0)
    func_aux(ecrit, 'ECRIT', 'rang', 'Can _cod')


    ##### MP ORAL #####
    oral = pd.read_excel(path/'Oral_MP.xlsx', header=0)
    func_aux(oral, 'ORAL', 'rang', 'Can _cod')

    ##### PC ORAL #####
    oral = pd.read_excel(path/'Oral_PC.xlsx', header=0)
    func_aux(oral, 'ORAL', 'rang', 'Can _cod')

    ##### PSI ORAL #####
    oral = pd.read_excel(path/'Oral_PSI.xlsx', header=0)
    func_aux(oral, 'ORAL', 'rang', 'Can _cod')

    ##### PT ORAL #####
    oral = pd.read_excel(path/'Oral_PT.xlsx', header=0)
    func_aux(oral, 'ORAL', 'rang', 'Can _cod')

    ##### TSI ORAL #####
    oral = pd.read_excel(path/'Oral_TSI.xlsx', header=0)
    func_aux(oral, 'ORAL', 'rang', 'Can _cod')


    ##### MP classes CMT #####
    donnees = pd.read_excel(path/'Classes_MP_CMT_spe_XXXX.xlsx', header=1)
    func_aux(donnees, 'RANG_ADMISSIBLE', 'rang_admissible', 'login')
    func_aux(donnees, 'RANG_CLASSE', 'rang_classe', 'login')

    ##### PC classes CMT #####
    donnees = pd.read_excel(path/'Classes_PC_CMT_spe_XXXX.xlsx', header=1)
    func_aux(donnees, 'RANG_ADMISSIBLE', 'rang_admissible', 'login')
    func_aux(donnees, 'RANG_CLASSE', 'rang_classe', 'login')

    ##### PSI classes CMT #####
    donnees = pd.read_excel(path/'Classes_PSI_CMT_spe_XXXX.xlsx', header=1)
    func_aux(donnees, 'RANG_ADMISSIBLE', 'rang_admissible', 'login')
    func_aux(donnees, 'RANG_CLASSE', 'rang_classe', 'login')

    ##### PT classes CMT #####
    donnees = pd.read_excel(path/'Classes_PT_CMT_spe_XXXX.xlsx', header=1)
    func_aux(donnees, 'RANG_ADMISSIBLE', 'rang_admissible', 'login')
    func_aux(donnees, 'RANG_CLASSE', 'rang_classe', 'login')

    ##### TSI classes CMT #####
    donnees = pd.read_excel(path/'Classes_TSI_CMT_spe_XXXX.xlsx', header=1)
    func_aux(donnees, 'RANG_ADMISSIBLE', 'rang_admissible', 'login')
    func_aux(donnees, 'RANG_CLASSE', 'rang_classe', 'login')

    ##### ATS
    data = pd.read_excel(path/'ResultatEcrit_DD_MM_YYYY_ATS.xlsx', header=0, skiprows= (1,))
    func_aux(data, 'ECRIT', 'Rang', "Numéro d'inscription")

    data = pd.read_excel(path / 'ResultatOral_DD_MM_YYYY_ATS.xlsx', header=0, skiprows= (1,))
    func_aux(data, 'ORAL', 'Rang', "Numéro d'inscription")

    data = pd.read_excel(path / 'ADMISSIBLE_ATS.xlsx', header=0)
    func_aux(data, 'RANG_ADMISSIBLE', 'rang', "Can _cod")


def extrait_etat_classes_type_admissible(path: Path):
    dico_etat_classes = {}
    dico_type_admissible = {}

    for f in ('MP', 'PC', 'PSI', 'PT', 'TSI'):
        data = pd.read_csv(path/f'Classes_{f}_CMT_spe_XXXX_SCEI.csv', sep = ";", header = 0)
        for i in range(len(data)):
            if repr(data['scei'][i]) != 'nan' and repr(data['etat'][i]) != 'nan':
                dico_etat_classes[int(data['scei'][i])] = int(data['etat'][i])

        data = pd.read_excel(path/f'Classes_{f}_CMT_spe_XXXX.xlsx', header=1)
        for i in range(len(data)):
            if repr(data['login'][i]) != 'nan' and repr(data['type_admissible'][i]) != 'nan':
                dico_type_admissible[int(data['login'][i])] = data['type_admissible'][i]

    return dico_etat_classes, dico_type_admissible



if __name__ == '__main__':
    path_files = Path("../PPII_ressources/data/public")
    db = sql.connect("concours.db")
    with db:
        print('epreuve')
        import_epreuves(db)
        print('voeux')
        import_voeux(db,path_files)
        print('ecole')
        import_ecole(db,path_files)
        print('etat reponse')
        import_etat_reponse(db,path_files)
        print('etablissement')
        import_etablissement(db,path_files)
        print('notes')
        import_notes(db,path_files)
        print('rang')
        import_rang(db,path_files)
        print('inscription')
        import_inscription(db,path_files)
        print('ats')
        import_ats(db,path_files)
    db.close()
else:
    import click
    @click.command()
    @click.argument("database", type=click.Path(dir_okay=False))
    @click.argument("src_dir", type=click.Path(exists=True, file_okay=False))
    @click.option("--init", "-i", default=None, type=click.Path(exists=True, dir_okay=False))
    def cli(database, src_dir, init):
        path_db = Path(database)
        path_files = Path(src_dir)
        db = sql.connect(path_db)
        if init is not None:
            init = Path(init)
            click.echo(f"Initialisation de {path_db.name} avec {init.name}")
            with open(init, "r") as f_init:
                with db:
                    db.executescript(f_init.read())

        click.echo("Importation des fichiers")
        with db:
            click.echo('epreuve')
            import_epreuves(db)
            click.echo('voeux')
            import_voeux(db, path_files)
            click.echo('ecole')
            import_ecole(db, path_files)
            click.echo('etat reponse')
            import_etat_reponse(db, path_files)
            click.echo('etablissement')
            import_etablissement(db, path_files)
            click.echo('notes')
            import_notes(db, path_files)
            click.echo('rang')
            import_rang(db, path_files)
            click.echo('inscription')
            import_inscription(db, path_files)
            click.echo('ats')
            import_ats(db, path_files)
        db.close()
