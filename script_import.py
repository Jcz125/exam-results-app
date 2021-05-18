import openpyxl as xl
import sqlite3 as sql
from pathlib import Path


def import_voeux():
    for classe in ("MP", "TSI", "PT", "ATS", "PC", "PSI"):
        file = xl.load_workbook(Path(f"../PPII_ressources/data/public/listeVoeux_{classe}.xlsx"), read_only=True)
        tab = file[file.sheetnames[0]]

        with db:
            c = db.cursor()
            rows = tab.rows
            next(rows)
            for line in rows:
                c.execute("INSERT OR IGNORE INTO voeux VALUES (?,?,?)", (line[0].value, line[2].value, line[3].value))
        file.close()


def import_candidat():
    file = xl.load_workbook(Path(f"../PPII_ressources/data/public/Inscription.xlsx"), read_only=True)
    tab = file[file.sheetnames[0]]

    with db:
        rows = tab.rows
        next(rows)
        C = {case.value: case.column - 1 for case in next(rows)}
        for line in rows:

            date_n = '-'.join(reversed(line[C['DATE_NAISSANCE']].value.split('/')))

            c = db.cursor()
            c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',(line[C['EPREUVE1']].value,line[C['OPTION1']].value))
            o1 = c.fetchone()
            if o1 is not None:
                o1 = o1[0]

            c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',(line[C['EPREUVE2']].value, line[C['OPTION2']].value))
            o2 = c.fetchone()
            if o2 is not None:
                o2 = o2[0]

            c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',
                (line[C['EPREUVE3']].value, line[C['OPTION3']].value))
            o3 = c.fetchone()
            if o3 is not None:
                o3 = o3[0]

            c.execute('SELECT id FROM ep_option WHERE (epreuve,option)=(?,?)',
                (line[C['EPREUVE4']].value, line[C['OPTION4']].value))
            o4 = c.fetchone()
            if o4 is not None:
                o4 = o4[0]

            if line[C['DECLARATION_HANDICAP']].value == 'non':
                handi = 0
            else:
                handi = 1

            c.execute('SELECT code FROM classe WHERE lib = ?',(line[C['CLASSE']].value,))
            classe = c.fetchone()
            if classe is None:
                c.execute('INSERT INTO classe(lib) VALUES(?)', (line[C['CLASSE']].value,))
                c.execute('SELECT code FROM classe WHERE lib = ?', (line[C['CLASSE']].value,))
                classe = c.fetchone()
            classe = classe[0]

            c.execute(
                "INSERT INTO candidat VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (line[C['CODE_CANDIDAT']].value, line[C['CIVILITE']].value, line[C['NOM']].value, line[C['PRENOM']].value, date_n,
                classe, line[C['PUISSANCE']].value, int(line[C['CODE_CONCOURS']].value), line[C['CODE_ETABLISSEMENT']].value, line[C['ADRESSE1']].value,
                line[C['ADRESSE2']].value, line[C['CP']].value, line[C['VILLE']].value, line[C['CODE_PAYS']].value, line[C['EMAIL']].value,
                line[C['TELEPHONE']].value, line[C['TEL_PORTABLE']].value, line[C['CODE_PAYS_NAISSANCE']].value, line[C['CODE_PAYS_NATIONALITE']].value, line[C['LIBELLE_VILLE_ECRIT']].value,
                line[C['NUMERO_INE']].value, int(line[C['COD_CSP_PERE']].value), int(line[C['COD_CSP_MERE']].value), int(line[C['CODE_SERIE']].value), int(line[C['ANNEE_BAC']].value),
                int(line[C['MOIS_BAC']].value), line[C['MENTION']].value, line[C['CAN_DEP_BAC']].value, int(line[C['CODE_ETAT_DOSSIER']].value), line[C['QUALITE']].value,
                handi, line[C['ARRONDISSEMENT_NAISSANCE']].value, o1, o2, o3, o4, line[C['SUJET_TIPE']].value, None, None))

    file.close()


def import_ats():
    file = xl.load_workbook(Path(f"../PPII_ressources/data/public/ADMISSIBLE_ATS.xlsx"), read_only=True)
    tab = file[file.sheetnames[0]]

    with db:
        c = db.cursor()
        rows = tab.rows
        C = {case.value: case.column - 1 for case in next(rows)}

        c.execute("INSERT INTO classe(lib) VALUES('ATS')")
        classe = c.execute("SELECT code FROM classe WHERE lib = 'ATS'").fetchone()
        classe = classe[0]

        for line in rows:
            if line[C['Civ _lib']].value == 'M.':
                civ = 1
            else:
                civ = 2

            pays = c.execute('SELECT code FROM pays WHERE lib = ?', (line[C['Can _pay _adr']].value,)).fetchone()
            if pays is None:
                c.execute('INSERT INTO pays(lib) VALUES(?)', (line[C['Can _pay _adr']].value,))
                pays = c.execute('SELECT code FROM pays WHERE lib = ?', (line[C['Can _pay _adr']].value,)).fetchone()
            pays = pays[0]

            c.execute(
                "INSERT INTO candidat(code, civ, nom, prenom, adresse1, adresse2, code_postal, commune, code_adr_pays, mail, tel, por, classe) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (line[C['Can _cod']].value, civ, line[C['Nom']].value, line[C['Prenom']].value, line[C['Can _ad 1']].value, line[C['Can _ad 2']].value, line[C['Can _cod _pos']].value, line[C['Can _com']].value, pays, line[C['Can _mel']].value, line[C['Can _tel']].value, line[C['Can _por']].value, classe))
    file.close()


if __name__ == '__main__':
    db = sql.connect("concours.db")
    import_voeux()
    import_candidat()
    import_ats()
    db.close()
